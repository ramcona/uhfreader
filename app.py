import sys
import os
import threading
import webbrowser
import time
import socket
from contextlib import closing
from openpyxl import Workbook, load_workbook

# Check if we're running as a PyInstaller bundle
if getattr(sys, 'frozen', False):
    # Running as a PyInstaller bundle
    application_path = os.path.dirname(sys.executable)
else:
    # Running as a regular Python script
    application_path = os.path.dirname(os.path.abspath(__file__))

# Add the application path to sys.path
sys.path.insert(0, application_path)

# Import Flask app components
from flask import Flask, render_template, request, send_file, Response, jsonify
from rfid_reader import RFIDReader
import json

# Find an available port
def find_free_port():
    with closing(socket.socket(socket.AF_INET, socket.SOCK_STREAM)) as s:
        s.bind(('', 0))
        s.setsockopt(socket.SOL_SOCKET, socket.SO_REUSEADDR, 1)
        return s.getsockname()[1]

# Initialize Flask app
app = Flask(__name__)
app.config['SECRET_KEY'] = 'your-secret-key-here'

# Initialize RFID reader
reader = RFIDReader()

@app.route('/')
def index():
    return render_template('index.html', 
                         settings=reader.settings,
                         data=reader.get_data(),
                         ports=reader.list_serial_ports())

@app.route('/get_ports')
def get_ports():
    """Fetch the list of serial ports and their status."""
    ports = reader.list_serial_ports()
    return jsonify(ports)

@app.route('/update_settings', methods=['POST'])
def update_settings():
    port = request.form['serial_port']
    baud_rate = int(request.form['baud_rate'])
    output_file = request.form['output_file']
    
    success, message = reader.setup_connection(port, baud_rate)
    if success:
        reader.settings['output_file'] = output_file
    return jsonify({'success': success, 'message': message})

@app.route('/start_reader')
def start_reader():
    success, message = reader.start()
    return jsonify({'success': success, 'message': message})

@app.route('/stop_reader')
def stop_reader():
    success, message = reader.stop()
    return jsonify({'success': success, 'message': message})

@app.route('/clear_data')
def clear_data():
    reader.clear_data()  # Call the clear_data method
    return jsonify({'success': True, 'message': 'Data cleared successfully'})

@app.route('/download_csv')
def download_csv():
    return send_file(reader.settings['csv_file'], as_attachment=True)

@app.route('/download_excel')
def download_excel():
    return send_file(reader.settings['output_file'], as_attachment=True)

@app.route('/import_participants', methods=['POST'])
def import_participants():
    if 'file' not in request.files:
        return jsonify({'success': False, 'message': 'No file uploaded'})
    file = request.files['file']
    if file.filename == '':
        return jsonify({'success': False, 'message': 'No file selected'})
    file_path = os.path.join(application_path, file.filename)
    file.save(file_path)
    reader.import_participants(file_path)
    return jsonify({'success': True, 'message': 'Participant data imported successfully'})

@app.route('/download_start')
def download_start():
    if not os.path.exists(reader.settings['output_file']):
        return jsonify({'success': False, 'message': 'File does not exist'}), 404
    return send_file(reader.settings['output_file'], as_attachment=True, download_name="start_data.xlsx")

@app.route('/download_finish')
def download_finish():
    if not os.path.exists(reader.settings['output_file']):
        return jsonify({'success': False, 'message': 'File does not exist'}), 404
    return send_file(reader.settings['output_file'], as_attachment=True, download_name="finish_data.xlsx")

@app.route('/download_merged')
def download_merged():
    if not os.path.exists(reader.settings['output_file']):
        return jsonify({'success': False, 'message': 'File does not exist'}), 404

    # Fetch merged data
    merged_data = reader.get_merged_data()

    # Create a new workbook and add headers
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["EPC", "NAMA", "BIB", "Start Timestamp", "Finish Timestamp", "Duration"])

    # Add merged data to the sheet
    for row in merged_data:
        sheet.append([
            row["EPC"],
            row["NAMA"],
            row["BIB"],
            row["Start Timestamp"],
            row["Finish Timestamp"],
            row["Duration"]
        ])

    # Save the workbook to a temporary file
    file_path = os.path.join(application_path, "merged_data.xlsx")
    workbook.save(file_path)

    # Send the file as a response
    return send_file(file_path, as_attachment=True, download_name="merged_data.xlsx")

@app.route('/retry_missed_tags')
def retry_missed_tags():
    success = reader.retry_missed_tags()
    return jsonify({'success': success, 'message': 'Retry operation completed'})

@app.route('/get_participants')
def get_participants():
    try:
        if not os.path.exists(reader.settings['output_file']):
            return jsonify({'success': False, 'message': 'File does not exist', 'data': []})

        workbook = load_workbook(reader.settings['output_file'])
        if "Participants" not in workbook.sheetnames:
            return jsonify({'success': False, 'message': 'Participants sheet does not exist', 'data': []})

        sheet = workbook["Participants"]
        participants = []
        for row in sheet.iter_rows(min_row=2, values_only=True):  # Skip header row
            participants.append({
                "Member NO": row[0],
                "Nama": row[1],
                "Alamat": row[2],
                "Gender": row[3],
                "EPC": row[4],
                "Country": row[5],
                "Status": row[6]
            })
        return jsonify({'success': True, 'data': participants})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e), 'data': []})

@app.route('/stream')
def stream():
    def event_stream():
        while True:
            data = json.dumps({
                'data': reader.get_data(),
                'stats': reader.get_stats(),
                'is_running': reader.running  # Add reader status
            })
            yield f"data: {data}\n\n"
            threading.Event().wait(0.5)
    return Response(event_stream(), mimetype="text/event-stream")

def run_flask_app(port):
    app.run(host='127.0.0.1', port=port, debug=False)

def main():
    # Find an available port
    port = find_free_port()
    
    # Start the Flask app in a separate thread
    flask_thread = threading.Thread(target=run_flask_app, args=(port,))
    flask_thread.daemon = True
    flask_thread.start()
    
    # Wait a moment for the Flask app to start
    time.sleep(1)
    
    # Open the web browser to the Flask app
    url = f"http://127.0.0.1:{port}"
    webbrowser.open(url)
    
    # Keep the main thread alive
    try:
        while True:
            time.sleep(1)
    except KeyboardInterrupt:
        print("Shutting down...")
        # Clean up resources
        if reader.running:
            reader.stop()
        sys.exit(0)

if __name__ == '__main__':
    main()