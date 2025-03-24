import serial
import csv
import time
import glob
from datetime import datetime
import threading
import serial.tools.list_ports
from threading import Thread
from queue import Queue
import os
from openpyxl import Workbook, load_workbook


class RFIDReader:
    def __init__(self):
        self.serial_conn = None
        self.current_data = []
        self.raw_packets = []
        self.settings = {
            'serial_port': None,
            'baud_rate': 57600,
            'output_file': 'rfid_data.xlsx'
        }
        self.lock = threading.Lock()
        self.running = False
        self.thread = None
        self.process_thread = None
        self.data_queue = Queue()  # Thread-safe queue for incoming data
        self.last_inventory_time = 0
        self.num_antennas = 0  # Will be dynamically set after querying the reader
        self.current_antenna = 1  # Track the current antenna port being used

    def query_antenna_ports(self):
        """Query the RFID reader to determine the number of supported antenna ports."""
        if self.serial_conn and self.serial_conn.is_open:
            try:
                # Send a command to query the number of antenna ports
                query_cmd = bytearray([0xA0, 0x02, 0x80, 0x22])  # Example command (adjust for your reader)
                self.serial_conn.write(query_cmd)
                time.sleep(0.1)  # Wait for the response

                # Read the response (adjust based on your reader's protocol)
                response = self.serial_conn.read(10)  # Adjust the number of bytes to read
                if len(response) >= 4:  # Ensure we have a valid response
                    self.num_antennas = response[3]  # Extract the number of antenna ports
                    print(f"Detected {self.num_antennas} antenna ports on the reader.")
                    return True
                else:
                    print("Failed to query antenna ports. Using default (4 ports).")
                    self.num_antennas = 4  # Fallback to a default value
                    return False
            except Exception as e:
                print(f"Error querying antenna ports: {e}")
                self.num_antennas = 4  # Fallback to a default value
                return False
        return False
        
    def list_serial_ports(self):
        """List all available serial ports and their connection status."""
        import serial.tools.list_ports
        import platform
        
        ports = []
        system = platform.system()
        
        if system == 'Windows':
            # Special handling for Windows
            for i in range(256):
                port_name = f'COM{i}'
                try:
                    # Just try to initialize the port without opening it
                    s = serial.Serial()
                    s.port = port_name
                    s.close()
                    
                    # If we get here, the port exists in the system
                    ports.append({
                        'device': port_name,
                        'description': f'COM Port {i}',
                        'status': 'Available'
                    })
                except serial.SerialException:
                    # Skip this port and continue to the next
                    continue
        else:
            # Original approach for macOS/Linux
            for port in serial.tools.list_ports.comports():
                try:
                    # Don't actually open the port, just check if it's listed
                    status = "Available"
                    ports.append({
                        'device': port.device,
                        'description': port.description,
                        'status': status
                    })
                except Exception:
                    pass
        
        # If still empty on Windows, try a more direct approach with list_ports
        if system == 'Windows' and len(ports) == 0:
            # Use the raw output from list_ports directly
            all_ports = list(serial.tools.list_ports.comports())
            for port in all_ports:
                ports.append({
                    'device': port.device,
                    'description': port.description,
                    'status': 'Unknown'
                })
        
        return ports
    
    def setup_connection(self, port, baud_rate=57600):
        """Establish a serial connection with the UHF reader."""
        try:
            self.serial_conn = serial.Serial(
                port=port,
                baudrate=baud_rate,
                parity=serial.PARITY_NONE,
                stopbits=serial.STOPBITS_ONE,
                bytesize=serial.EIGHTBITS,
                timeout=2,  # Increase timeout to 2 seconds
                write_timeout=2,  # Set write timeout
                rtscts=True,  # Enable hardware flow control (if supported)
                dsrdtr=True  # Enable hardware flow control (if supported)
            )
            self.settings['serial_port'] = port
            self.settings['baud_rate'] = baud_rate
            return True, f"Successfully connected to {port}"
        except serial.SerialException as e:
            return False, f"Error connecting to {port}: {e}"
    
    def stop(self):
        """Stop the RFID reader and processing threads."""
        self.running = False
        
        # Stop inventory if active for Chaofan reader
        if self.serial_conn and self.serial_conn.is_open:
            try:
                # Using the Chaofan command to stop inventory
                stop_cmd = bytearray([0xA0, 0x03, 0x00, 0xA3])
                self.serial_conn.write(stop_cmd)
                time.sleep(0.1)
            except Exception as e:
                print(f"Error sending stop command: {e}")
        
        if self.thread:
            self.thread.join(timeout=2)  # Wait for the read_loop thread to finish with timeout
        if self.process_thread:
            self.process_thread.join(timeout=2)  # Wait for the process_queue thread to finish with timeout
        if self.serial_conn:
            self.serial_conn.close()  # Close the serial connection
        return True, "Reader stopped"
    
    def write_to_excel(self, tag_data, sheet_name):
        """Write tag data to a specific sheet in the Excel file."""
        try:
            if os.path.exists(self.settings['output_file']):
                workbook = load_workbook(self.settings['output_file'])
            else:
                workbook = Workbook()
                # Remove the default sheet and create Start, Finish, and Participants sheets
                workbook.remove(workbook.active)
                workbook.create_sheet("Start")
                workbook.create_sheet("Finish")
                workbook.create_sheet("Participants")

            sheet = workbook[sheet_name]
            if sheet.max_row == 1:  # Add headers if the sheet is empty
                if sheet_name in ["Start", "Finish"]:
                    sheet.append(["EPC", "Timestamp", "Gate"])
                elif sheet_name == "Participants":
                    sheet.append(["Member NO", "Nama", "Alamat", "Gender", "EPC", "Country", "Status"])

            if sheet_name in ["Start", "Finish"]:
                sheet.append([tag_data['epc'], tag_data['timestamp'], sheet_name])
            workbook.save(self.settings['output_file'])
            print(f"Data saved to {sheet_name} sheet in {self.settings['output_file']}")
        except Exception as e:
            print(f"Error saving to Excel file: {e}")

    def clear_data(self, sheet_name=None):
        """Clear data from a specific sheet or all sheets."""
        try:
            if not os.path.exists(self.settings['output_file']):
                print(f"File {self.settings['output_file']} does not exist.")
                return

            workbook = load_workbook(self.settings['output_file'])
            if sheet_name:
                if sheet_name not in workbook.sheetnames:
                    print(f"Sheet {sheet_name} does not exist.")
                    return
                sheet = workbook[sheet_name]
                sheet.delete_rows(2, sheet.max_row)  # Delete all rows except headers
            else:
                for sheet_name in ["Start", "Finish", "Participants"]:
                    if sheet_name in workbook.sheetnames:
                        sheet = workbook[sheet_name]
                        sheet.delete_rows(2, sheet.max_row)
            workbook.save(self.settings['output_file'])
            print(f"Data cleared from {sheet_name if sheet_name else 'all sheets'}.")
        except Exception as e:
            print(f"Error clearing Excel file: {e}")

    def import_participants(self, file_path):
        """Import participant data from an Excel file."""
        try:
            if not os.path.exists(file_path):
                print(f"File {file_path} does not exist.")
                return

            workbook = load_workbook(file_path)
            sheet = workbook.active
            participants = []
            for row in sheet.iter_rows(min_row=2, values_only=True):
                participants.append({
                    "Member NO": row[0],
                    "Nama": row[1],
                    "Alamat": row[2],
                    "Gender": row[3],
                    "EPC": row[4],
                    "Country": row[5],
                    "Status": row[6]
                })

            if not os.path.exists(self.settings['output_file']):
                # Create a new workbook if it doesn't exist
                workbook = Workbook()
                workbook.remove(workbook.active)  # Remove default sheet
                workbook.create_sheet("Start")
                workbook.create_sheet("Finish")
                workbook.create_sheet("Participants")
                workbook.save(self.settings['output_file'])

            workbook = load_workbook(self.settings['output_file'])
            if "Participants" not in workbook.sheetnames:
                workbook.create_sheet("Participants")
            sheet = workbook["Participants"]
            sheet.delete_rows(2, sheet.max_row)  # Clear existing data
            for participant in participants:
                sheet.append(list(participant.values()))
            workbook.save(self.settings['output_file'])
            print("Participant data imported successfully.")
        except Exception as e:
            print(f"Error importing participant data: {e}")

    def get_merged_data(self):
        """Merge data from Start, Finish, and Participants sheets."""
        try:
            if not os.path.exists(self.settings['output_file']):
                return []

            workbook = load_workbook(self.settings['output_file'])
            start_sheet = workbook["Start"]
            finish_sheet = workbook["Finish"]
            participants_sheet = workbook["Participants"]

            # Create a dictionary of participants by EPC
            participants = {}
            for row in participants_sheet.iter_rows(min_row=2, values_only=True):
                participants[row[4]] = row  # EPC is the key

            # Create dictionaries for start and finish timestamps by EPC
            start_timestamps = {}
            for row in start_sheet.iter_rows(min_row=2, values_only=True):
                epc = row[0]
                timestamp = row[1]
                start_timestamps[epc] = timestamp

            finish_timestamps = {}
            for row in finish_sheet.iter_rows(min_row=2, values_only=True):
                epc = row[0]
                timestamp = row[1]
                finish_timestamps[epc] = timestamp

            # Merge data
            merged_data = []
            for epc in participants:
                nama = participants[epc][1]
                bib = participants[epc][0]
                start_time = start_timestamps.get(epc)
                finish_time = finish_timestamps.get(epc)

                # Calculate duration if both start and finish times are available
                duration = ""
                if start_time and finish_time:
                    start_dt = datetime.strptime(start_time, "%Y-%m-%d %H:%M:%S")
                    finish_dt = datetime.strptime(finish_time, "%Y-%m-%d %H:%M:%S")
                    delta = finish_dt - start_dt
                    duration = str(delta)  # Format as HH:mm:ss

                merged_data.append({
                    "EPC": epc,
                    "NAMA": nama,
                    "BIB": bib,
                    "Start Timestamp": start_time if start_time else "N/A",
                    "Finish Timestamp": finish_time if finish_time else "N/A",
                    "Duration": duration
                })

            return merged_data
        except Exception as e:
            print(f"Error merging data: {e}")
            return []
    
    def read_loop(self):
        """Main reading loop with improved buffer handling for Chaofan reader."""
        buffer = bytearray()
        while self.running:
            try:
                if self.serial_conn and self.serial_conn.in_waiting:
                    # Read all available data
                    new_data = self.serial_conn.read(self.serial_conn.in_waiting)
                    if new_data:
                        # For debugging
                        # print(f"Received data: {' '.join([f'{b:02X}' for b in new_data])}")
                        self.data_queue.put(new_data)
                
                # Periodically restart inventory to improve tag detection
                current_time = time.time()
                if current_time - self.last_inventory_time > 3:  # Every 3 seconds
                    self.start_fast_inventory()
                    self.last_inventory_time = current_time
                
                time.sleep(0.01)  # Small delay to prevent CPU hogging
            except Exception as e:
                print(f"Error in read loop: {e}")
                time.sleep(0.1)  # Longer delay after error

    def process_queue(self):
        """Process data from the queue."""
        buffer = bytearray()  # Buffer to accumulate data across queue entries
        
        while self.running:
            try:
                if not self.data_queue.empty():
                    data = self.data_queue.get()
                    
                    # Add to our accumulating buffer
                    buffer.extend(data)
                    
                    # Process complete packets from the buffer
                    i = 0
                    while i < len(buffer):
                        # Check for Chaofan packet structure
                        if len(buffer) - i >= 21:
                            # For debugging
                            # print(f"Processing potential packet: {' '.join([f'{b:02X}' for b in buffer[i:i+21]])}")
                            # Process the packet and check if it was valid
                            if self.process_data(buffer[i:i+21]):
                                i += 21  # Move past this packet
                            else:
                                i += 1   # Not a valid packet, move one byte
                        else:
                            # Not enough data for a complete packet
                            break
                    
                    # Keep only unprocessed data in the buffer
                    if i > 0:
                        buffer = buffer[i:]
                
                else:
                    time.sleep(0.01)  # Small delay when queue is empty
            except Exception as e:
                print(f"Error processing data from queue: {e}")
                time.sleep(0.1)  # Longer delay after error

    def start(self):
        """Start the RFID reader and processing threads."""
        if not self.serial_conn or not self.serial_conn.is_open:
            return False, "Serial connection not established"
        
        # Query the reader for the number of antenna ports
        self.query_antenna_ports()

        self.running = True
        
        # Start the reading thread
        self.thread = threading.Thread(target=self.read_loop, daemon=True)
        self.thread.start()

        # Start a separate thread for processing data
        self.process_thread = threading.Thread(target=self.process_queue, daemon=True)
        self.process_thread.start()
        
        # Start initial inventory
        self.start_fast_inventory()
        self.last_inventory_time = time.time()
        
        return True, "Reader started"
    
    # def process_data(self, data):
    #     """Process Chaofan UHF reader data packet and extract EPC."""
    #     # Check if this looks like a valid Chaofan packet
    #     if len(data) == 21:  # Expected length for Chaofan tag data
    #         try:
    #             # Chaofan packet format check (this may need adjustment based on your specific reader)
    #             epc = data[6:18]  # Extract EPC (12 bytes)
    #             epc_hex = ''.join([f"{b:02X}" for b in epc])  # Convert to hex string

    #             rssi = data[18]  # Extract RSSI
    #             antenna = data[19]  # Extract antenna number

    #             tag_data = {
    #                 'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S.%f'),
    #                 'epc': epc_hex,
    #                 'rssi': rssi,
    #                 'antenna': antenna,
    #                 'detected_as': 'chaofan_custom'
    #             }

    #             with self.lock:
    #                 # Check if this is a new tag
    #                 if not any(tag['epc'] == epc_hex for tag in self.current_data):
    #                     self.current_data.append(tag_data)
    #                     self.write_to_csv(tag_data)
    #                     print(f"Tag found: {epc_hex}, RSSI: {rssi}, Antenna: {antenna}")
    #                     print(f"Total tags processed: {len(self.current_data)}")
    #             return True  # Valid packet processed
    #         except Exception as e:
    #             print(f"Error parsing tag data: {e}")
    #             return False  # Error processing packet
    #     return False  # Not a valid packet
    
    def process_data(self, data):
        """Process Chaofan UHF reader data packet and extract EPC and antenna port information."""
        # Check if this looks like a valid Chaofan packet
        if len(data) == 21:  # Expected length for Chaofan tag data
            try:
                # Extract EPC (12 bytes)
                epc = data[6:18]
                epc_hex = ''.join([f"{b:02X}" for b in epc])  # Convert to hex string

                # Extract RSSI and antenna port number
                rssi = data[18]
                antenna_port = data[19]  # Antenna port number (1-N)

                # Validate antenna port number
                if antenna_port < 1 or antenna_port > self.num_antennas:
                    print(f"Invalid antenna port detected: {antenna_port}")
                    return False

                tag_data = {
                    'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S.%f'),
                    'epc': epc_hex,
                    'rssi': rssi,
                    'antenna_port': antenna_port,
                    'detected_as': 'chafon'
                }

                with self.lock:
                    # Check if this is a new tag
                    if not any(tag['epc'] == epc_hex for tag in self.current_data):
                        self.current_data.append(tag_data)
                        self.write_to_csv(tag_data)
                        print(f"Tag found: {epc_hex}, RSSI: {rssi}, Antenna Port: {antenna_port}")
                        print(f"Total tags processed: {len(self.current_data)}")
                return True  # Valid packet processed
            except Exception as e:
                print(f"Error parsing tag data: {e}")
                return False  # Error processing packet
        return False  # Not a valid packet

    def debug_print_bytes(self, data):
        """Print byte data in a readable format."""
        hex_str = ' '.join([f"{b:02X}" for b in data])
        return hex_str
    
    def write_to_csv(self, tag_data):
        """Write tag data to CSV."""
        try:
            with open(self.settings['csv_file'], 'a', newline='') as csvfile:
                fieldnames = ['timestamp', 'epc', 'rssi', 'antenna', 'detected_as']
                writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
                if csvfile.tell() == 0:
                    writer.writeheader()
                writer.writerow(tag_data)
        except Exception as e:
            print(f"Error saving to CSV: {e}")

    def get_data(self):
        """Get current tag data."""
        with self.lock:
            return self.current_data.copy()

    def get_stats(self):
        """Get statistics about the current session."""
        with self.lock:
            return {
                'total_reads': len(self.current_data),
                'last_read': self.current_data[0]['timestamp'] if self.current_data else 'Never'
            }
    
    def start_fast_inventory(self):
        """Send command to start fast inventory mode for Chaofan reader with all antenna ports."""
        if self.serial_conn and self.serial_conn.is_open:
            try:
                # Stop any ongoing inventory
                stop_cmd = bytearray([0xA0, 0x03, 0x00, 0xA3])
                self.serial_conn.write(stop_cmd)
                time.sleep(0.1)

                # Configure the reader to use all antenna ports (1 to num_antennas)
                antenna_config_cmd = bytearray([0xA0, 0x0B, 0x00, self.num_antennas])
                for i in range(1, self.num_antennas + 1):
                    antenna_config_cmd.append(i)  # Add antenna ports (1, 2, 3, ..., N)
                antenna_config_cmd.extend([0x00] * (8 - self.num_antennas))  # Pad with zeros if necessary
                antenna_config_cmd.append(0x00)  # Checksum placeholder (adjust as needed)
                self.serial_conn.write(antenna_config_cmd)
                time.sleep(0.1)

                # Start fast inventory with all antenna ports
                fast_inventory_cmd = bytearray([0xA0, 0x06, 0x01, 0xFF, 0x10, 0x20, 0xD6])
                self.serial_conn.write(fast_inventory_cmd)

                print(f"Fast inventory started with {self.num_antennas} antenna ports.")
                return True
            except Exception as e:
                print(f"Error starting fast inventory with antenna ports: {e}")
                return False
        return False
    
    def retry_missed_tags(self):
        """Attempt to detect any potentially missed tags by restarting inventory with different parameters."""
        if not self.serial_conn or not self.serial_conn.is_open:
            return False
            
        try:
            # Stop current inventory
            stop_cmd = bytearray([0xA0, 0x03, 0x00, 0xA3])
            self.serial_conn.write(stop_cmd)
            time.sleep(0.2)
            
            # Clear input buffer
            self.serial_conn.reset_input_buffer()
            
            # Set higher power (if supported by your reader)
            power_cmd = bytearray([0xA0, 0x07, 0x3B, 0x30, 0x00, 0x00, 0x12])
            self.serial_conn.write(power_cmd)
            time.sleep(0.2)
            
            # Start inventory with different parameters to catch missed tags
            # Using alternative antenna settings and sensitivity
            alt_inventory_cmd = bytearray([0xA0, 0x06, 0x01, 0xF0, 0x10, 0x10, 0xC7])
            self.serial_conn.write(alt_inventory_cmd)
            
            print("Retrying with alternate settings to detect missed tags")
            return True
        except Exception as e:
            print(f"Error in retry operation: {e}")
            return False